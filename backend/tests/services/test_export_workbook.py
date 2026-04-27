"""ExportService workbook: formula tokens and ticker metrics block."""

from __future__ import annotations

from datetime import UTC, date, datetime

import openpyxl
import pytest

from app.schemas import (
    AnalyticsPerformanceResult,
    CALPoint,
    CompletePortfolio,
    CorrelationMatrix,
    CovarianceMatrix,
    FamaFrenchThreePerTicker,
    FrontierPoint,
    MarketMetrics,
    ORP,
    ORPPerformanceMetrics,
    OptimizationResult,
    PriceBar,
    StockMetrics,
    TickerValuationBlock,
    ValuationResult,
)
from app.services.export_service import ExportService


def _monthly_dates(n: int, start: date) -> list[date]:
    out: list[date] = []
    y, m = start.year, start.month
    for _ in range(n):
        out.append(date(y, m, 1))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return out


def _price_bars(dates: list[date], seed: float) -> list[PriceBar]:
    return [
        PriceBar(
            date=d,
            open=100 + seed,
            high=101 + seed,
            low=99 + seed,
            close=100 + i * 0.001 + seed,
            volume=1,
        )
        for i, d in enumerate(dates)
    ]


@pytest.fixture
def minimal_export_inputs():
    n = 130
    dates = _monthly_dates(n, date(2010, 1, 1))
    bars_ibm = _price_bars(dates, 0.0)
    bars_aaoi = _price_bars(dates, 2.0)
    spy_bars = [
        PriceBar(
            date=d,
            open=400,
            high=401,
            low=399,
            close=400 + i * 0.001,
            volume=1,
        )
        for i, d in enumerate(dates)
    ]
    rf_series = [{"date": f"{d.year}-{d.month:02d}-01", "rate": 0.02 / 12.0} for d in dates]

    tickers = ["IBM", "AAOI"]
    stocks = [
        StockMetrics(
            ticker=t,
            expected_return=0.1,
            std_dev=0.2,
            beta=1.0,
            alpha=0.0,
            firm_specific_var=0.01,
            n_observations=n,
        )
        for t in tickers
    ]
    cov = CovarianceMatrix(
        tickers=tickers,
        matrix=[[0.04, 0.01], [0.01, 0.04]],
    )
    corr = CorrelationMatrix(
        tickers=tickers,
        matrix=[[1.0, 0.25], [0.25, 1.0]],
    )
    opt = OptimizationResult(
        request_id="opt_export_test",
        as_of=datetime(2025, 4, 1, tzinfo=UTC),
        risk_free_rate=0.05,
        market=MarketMetrics(expected_return=0.1, std_dev=0.2, variance=0.04),
        stocks=stocks,
        covariance=cov,
        correlation=corr,
        orp=ORP(
            weights={"IBM": 0.6, "AAOI": 0.4},
            expected_return=0.12,
            std_dev=0.2,
            variance=0.04,
            sharpe=0.5,
        ),
        complete=CompletePortfolio(
            y_star=0.8,
            weight_risk_free=0.2,
            weights={"IBM": 0.5, "AAOI": 0.3},
            expected_return=0.1,
            std_dev=0.15,
            leverage_used=False,
        ),
        frontier_points=[FrontierPoint(std_dev=0.1, expected_return=0.08)],
        cal_points=[CALPoint(std_dev=0.0, expected_return=0.05, y=0.0)],
        warnings=[],
    )
    ana = AnalyticsPerformanceResult(
        as_of=datetime.now(UTC),
        window_start=dates[0],
        window_end=dates[-1],
        risk_free_rate=0.05,
        data_source="test",
        orp=ORPPerformanceMetrics(
            treynor=0.1,
            jensen_alpha=0.02,
            n_observations=100,
            total_variance=0.1,
            systematic_variance=0.05,
            unsystematic_variance=0.05,
            sim_variance_mismatch=0.0,
        ),
        complete=None,
        holding=[],
        fama_french=[
            FamaFrenchThreePerTicker(
                ticker="IBM",
                beta_mkt=1.0,
                beta_smb=0.0,
                beta_hml=0.0,
                alpha=0.01,
                n_observations=100,
                expected_return_ff3=0.1,
                expected_return_capm=0.1,
            ),
            FamaFrenchThreePerTicker(
                ticker="AAOI",
                beta_mkt=1.0,
                beta_smb=0.0,
                beta_hml=0.0,
                alpha=0.01,
                n_observations=100,
                expected_return_ff3=0.1,
                expected_return_capm=0.1,
            ),
        ],
        market=MarketMetrics(expected_return=0.1, std_dev=0.2, variance=0.04),
        warnings=[],
    )

    def _vblock(t: str, bars: list[PriceBar]) -> TickerValuationBlock:
        return TickerValuationBlock(
            ticker=t,
            fcff=1e6,
            fcfe=1e6,
            fcff_value_per_share=100.0,
            fcfe_value_per_share=100.0,
            ddm_gordon=100.0,
            ddm_two_stage=100.0,
            cost_of_equity=0.1,
            wacc=0.09,
            sustainable_growth_rate=0.05,
            roe=0.15,
            gross_margin=0.4,
            book_value_per_share=50.0,
            earnings_per_share=5.0,
            price_to_book=2.0,
            price_to_earnings=20.0,
            historical_prices=bars,
            warnings=[],
        )

    val = ValuationResult(
        as_of=datetime.now(UTC),
        per_ticker=[_vblock("IBM", bars_ibm), _vblock("AAOI", bars_aaoi)],
        data_source="test",
        warnings=[],
    )
    return opt, ana, val, spy_bars, rf_series


def test_export_ticker_metrics_use_varp_stdevp_and_annualized_rows(minimal_export_inputs):
    opt, ana, val, spy_bars, rf_series = minimal_export_inputs
    bio = ExportService().build_workbook(opt, ana, val, spy_bars, rf_series)
    wb = openpyxl.load_workbook(bio, data_only=False)
    for name in ("IBM", "AAOI"):
        ws = wb[name]
        assert "VARP" in (ws["M4"].value or "")
        assert "STDEVP" in (ws["M5"].value or "")
        assert ws["M7"].value == "=(1+M2)^12-1"
        assert ws["M8"].value == "=M5*SQRT(12)"
        assert "_xludf" not in str(ws["M4"].value)


def test_export_portfolio_weight_risk_free_formula(minimal_export_inputs):
    opt, ana, val, spy_bars, rf_series = minimal_export_inputs
    bio = ExportService().build_workbook(opt, ana, val, spy_bars, rf_series)
    wb = openpyxl.load_workbook(bio, data_only=False)
    ws = wb["Portfolio Organization"]
    # y* row then w_rf = 1 - y*
    found = False
    for r in range(1, 30):
        if ws.cell(row=r, column=1).value == "Weight in Risky (y*)":
            assert str(ws.cell(row=r + 1, column=2).value or "").startswith("=1-B")
            found = True
            break
    assert found


def test_export_correlation_matrix_uses_correl(minimal_export_inputs):
    opt, ana, val, spy_bars, rf_series = minimal_export_inputs
    bio = ExportService().build_workbook(opt, ana, val, spy_bars, rf_series)
    wb = openpyxl.load_workbook(bio, data_only=False)
    ws = wb["Portfolio Organization"]
    found = False
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.startswith("=CORREL("):
                found = True
                break
        if found:
            break
    assert found
