"""ExportService workbook builders — formula coverage on key sheets."""

from __future__ import annotations

from datetime import UTC, datetime
from datetime import date as Date

import numpy as np
import openpyxl
import pandas as pd
import pytest

from app.schemas import (
    AnalyticsPerformanceResult,
    CompletePerformanceMetrics,
    ExportRequest,
    FamaFrenchThreePerTicker,
    HoldingPeriodMonthlyReturns,
    MarketMetrics,
    ORPPerformanceMetrics,
    PriceBar,
    RiskProfile,
    TickerValuationBlock,
    ValuationExportDrivers,
    ValuationResult,
)
from app.services.export_service import ExportService
from app.services.optimize_service import MARKET_PROXY_TICKER
from quant.returns import annualization_factor
from quant.types import (
    CALPoint,
    CompletePortfolio,
    CorrelationMatrix,
    CovarianceMatrix,
    FrontierPoint,
    MarketMetrics as QMarket,
    OptimizationResult,
    ORP,
    StockMetrics,
)


def _price_bars() -> list[PriceBar]:
    return [
        PriceBar(
            date=Date(2024, i, 1),
            open=100.0,
            high=101.0,
            low=99.0,
            close=100.0 + i * 0.1,
            close_nominal=100.0 + i * 0.1,
            volume=1_000_000,
        )
        for i in range(1, 6)
    ]


def _aligned_log_returns(n: int = 50) -> pd.DataFrame:
    rng = np.random.default_rng(42)
    idx = pd.date_range("2023-01-03", periods=n, freq="D")
    return pd.DataFrame(
        {
            "AAA": rng.normal(0.0004, 0.015, n),
            "BBB": rng.normal(0.0003, 0.014, n),
            MARKET_PROXY_TICKER: rng.normal(0.00035, 0.012, n),
        },
        index=idx,
    )


def _ff_monthly() -> pd.DataFrame:
    idx = pd.to_datetime(["2023-01-31", "2023-02-28", "2023-03-31", "2023-04-30"])
    return pd.DataFrame(
        {
            "RF": [0.001] * 4,
            "Mkt_RF": [0.01, -0.02, 0.015, 0.005],
            "SMB": [0.001, 0.002, -0.001, 0.0],
            "HML": [0.0, 0.001, -0.002, 0.001],
            "AAA": [0.02, -0.03, 0.01, 0.008],
            "BBB": [0.015, -0.025, 0.012, 0.007],
        },
        index=idx,
    )


def _optimization() -> OptimizationResult:
    tickers = ["AAA", "BBB"]
    stocks = [
        StockMetrics(
            ticker="AAA",
            expected_return=0.15,
            std_dev=0.25,
            beta=1.1,
            alpha=0.02,
            firm_specific_var=0.01,
            n_observations=500,
        ),
        StockMetrics(
            ticker="BBB",
            expected_return=0.12,
            std_dev=0.22,
            beta=0.9,
            alpha=0.01,
            firm_specific_var=0.015,
            n_observations=500,
        ),
    ]
    cov = CovarianceMatrix(
        tickers=tickers,
        matrix=[[0.0625, 0.03], [0.03, 0.0484]],
    )
    rho = [[1.0, 0.5], [0.5, 1.0]]
    cor = CorrelationMatrix(tickers=tickers, matrix=rho)
    orp = ORP(
        weights={"AAA": 0.6, "BBB": 0.4},
        expected_return=0.14,
        std_dev=0.2,
        variance=0.04,
        sharpe=0.65,
    )
    cp = CompletePortfolio(
        y_star=0.9,
        weight_risk_free=0.1,
        weights={"AAA": 0.54, "BBB": 0.36},
        expected_return=0.132,
        std_dev=0.18,
        leverage_used=False,
    )
    mkt = QMarket(expected_return=0.1, std_dev=0.18, variance=0.0324)

    return OptimizationResult(
        request_id="opt_test",
        as_of=datetime(2025, 1, 15, tzinfo=UTC),
        risk_free_rate=0.04,
        market=mkt,
        stocks=stocks,
        covariance=cov,
        correlation=cor,
        orp=orp,
        complete=cp,
        frontier_points=[FrontierPoint(std_dev=0.16, expected_return=0.11)],
        cal_points=[CALPoint(std_dev=0.0, expected_return=0.04, y=0.0)],
        warnings=[],
    )


def _analytics() -> AnalyticsPerformanceResult:
    z = 0.0
    zm = ORPPerformanceMetrics(
        treynor=0.03,
        jensen_alpha=0.015,
        n_observations=240,
        total_variance=z,
        systematic_variance=z,
        unsystematic_variance=z,
        sim_variance_mismatch=z,
    )
    fm = CompletePerformanceMetrics(
        treynor=0.02,
        jensen_alpha=0.01,
        n_observations=240,
        total_variance=z,
        systematic_variance=z,
        unsystematic_variance=z,
        sim_variance_mismatch=z,
    )
    m = MarketMetrics(expected_return=0.1, std_dev=0.18, variance=0.0324)
    ff = [
        FamaFrenchThreePerTicker(
            ticker="AAA",
            beta_mkt=1.0,
            beta_smb=0.2,
            beta_hml=0.1,
            alpha=0.012,
            n_observations=100,
            expected_return_ff3=0.11,
            expected_return_capm=0.12,
        ),
        FamaFrenchThreePerTicker(
            ticker="BBB",
            beta_mkt=0.95,
            beta_smb=-0.1,
            beta_hml=0.05,
            alpha=0.008,
            n_observations=100,
            expected_return_ff3=0.09,
            expected_return_capm=0.1,
        ),
    ]
    return AnalyticsPerformanceResult(
        as_of=datetime(2025, 1, 15, tzinfo=UTC),
        window_start=Date(2020, 1, 1),
        window_end=Date(2025, 1, 15),
        risk_free_rate=0.04,
        data_source="test",
        orp=zm,
        complete=fm,
        holding=[
            HoldingPeriodMonthlyReturns(
                years=5,
                n_observations=60,
                window_start=Date(2020, 1, 31),
                window_end=Date(2025, 1, 31),
                arithmetic_mean_monthly_return=0.01,
                geometric_mean_monthly_return=0.009,
            )
        ],
        fama_french=ff,
        market=m,
        warnings=[],
    )


def _valuation() -> ValuationResult:
    b1 = TickerValuationBlock(
        ticker="AAA",
        fcff=1_000_000.0,
        fcfe=800_000.0,
        fcff_value_per_share=120.5,
        fcfe_value_per_share=110.2,
        ddm_gordon=95.0,
        ddm_two_stage=100.0,
        cost_of_equity=0.1,
        wacc=0.09,
        sustainable_growth_rate=0.04,
        roe=0.15,
        gross_margin=0.4,
        operating_margin=0.2,
        price_to_earnings=18.5,
        price_to_book=3.2,
        earnings_per_share=5.0,
        book_value_per_share=29.5,
        warnings=[],
    )
    b2 = TickerValuationBlock(
        ticker="BBB",
        fcff=None,
        fcfe=None,
        fcff_value_per_share=None,
        fcfe_value_per_share=None,
        ddm_gordon=None,
        ddm_two_stage=None,
        cost_of_equity=0.11,
        wacc=0.1,
        warnings=[],
    )
    drivers = [
        ValuationExportDrivers(
            ticker="AAA",
            ebit=2_000_000.0,
            tax_rate=0.21,
            depreciation=500_000.0,
            capex=400_000.0,
            delta_nwc=50_000.0,
            interest_expense=100_000.0,
            net_borrowing=20_000.0,
            beta=1.1,
            risk_free_annual=0.04,
            market_cap=50_000_000.0,
            total_debt=10_000_000.0,
            cost_of_debt_pretax=0.05,
        ),
        ValuationExportDrivers(
            ticker="BBB",
            ebit=1_000_000.0,
            tax_rate=0.21,
            depreciation=200_000.0,
            capex=150_000.0,
            delta_nwc=10_000.0,
            interest_expense=80_000.0,
            net_borrowing=0.0,
            beta=0.95,
            risk_free_annual=0.04,
            market_cap=30_000_000.0,
            total_debt=15_000_000.0,
            cost_of_debt_pretax=0.055,
        ),
    ]
    return ValuationResult(
        as_of=datetime(2025, 1, 15, tzinfo=UTC),
        per_ticker=[b1, b2],
        data_source="test",
        warnings=[],
        export_drivers=drivers,
    )


@pytest.fixture
def export_ctx():
    svc = ExportService()
    rf_series = [{"date": "2020-01-01", "rate": 0.04}]
    body = ExportRequest(
        tickers=["AAA", "BBB"],
        risk_profile=RiskProfile(risk_aversion=3, target_return=None),
        allow_short=True,
        allow_leverage=True,
    )
    return svc, rf_series, body


def test_formula_sheets_have_formulas_where_expected(export_ctx):
    svc, rf_series, body = export_ctx
    optimize = _optimization()
    ana = _analytics()
    val = _valuation()

    frame = _aligned_log_returns()
    bio = svc.build_workbook(
        optimize=optimize,
        analytics=ana,
        valuation=val,
        market_index_bars=_price_bars(),
        rf_series=rf_series,
        risk_profile=RiskProfile(risk_aversion=3, target_return=None),
        allow_leverage=True,
        export_request=body,
        aligned_log_returns=frame,
        annualization_factor=annualization_factor(body.return_frequency),
        allow_short=body.allow_short,
        ff_monthly=_ff_monthly(),
    )

    wb = openpyxl.load_workbook(bio)

    oi = wb["Optimize Inputs"]
    orp_formula = str(oi.cell(row=9, column=6).value)
    assert "SUMPRODUCT(" in orp_formula
    assert orp_formula.startswith("=")

    aaa = wb["AAA"]
    for row in range(2, 7):
        for col in (13, 14, 15):
            c = aaa.cell(row=row, column=col)
            if c.data_type != "f":
                continue
            fv = c.value
            assert isinstance(fv, str)
            assert "@" not in fv
            assert "VAR.P((" not in fv
            assert "STDEV.P((" not in fv
            assert "AVERAGE((" not in fv
            assert "GEOMEAN((" not in fv

    pf = wb["Portfolio Organization"]
    for row in range(1, 60):
        label = pf.cell(row=row, column=1).value
        if label == "Risk-free rate r_f (linked to Optimize Inputs)":
            assert pf.cell(row=row, column=2).data_type == "f"
        if label == "Risk premium E(r_ORP) − r_f":
            assert pf.cell(row=row, column=2).data_type == "f"
        if label == "Final y* (target-return override + leverage rule)":
            assert pf.cell(row=row, column=2).data_type == "f"
        if label == "Variance ORP = w'Σw (sample cov × k)":
            vcell = pf.cell(row=row, column=2)
            assert vcell.data_type in {"n", "f"}

    weights_header = False
    in_weights = False
    for row in range(1, 60):
        if pf.cell(row=row, column=2).value == "w_ORP (Optimize Inputs)":
            in_weights = True
            continue
        if in_weights and pf.cell(row=row, column=1).value == "AAA":
            assert pf.cell(row=row, column=2).data_type == "f"
            weights_header = True
            break
    assert weights_header

    corr = False
    in_corr = False
    for row in range(1, 100):
        v1 = pf.cell(row=row, column=1).value
        if v1 == "Correlation matrix (from aligned log returns)":
            in_corr = True
            continue
        if in_corr and v1 == "AAA":
            assert pf.cell(row=row, column=2).data_type == "f"
            corr = True
            break
    assert corr

    tech = wb["Technical Analysis"]
    # Jensen display formula on same rows as backend inputs — find first Jensen row
    for row in range(1, 20):
        if tech.cell(row=row, column=1).value == "Jensen's Alpha (Annual)":
            assert tech.cell(row=row, column=3).data_type == "f"
            break
    else:  # pragma: no cover
        pytest.fail("Jensen row not found")

    val_ws = wb["Valuation and Earnings Analysis"]
    for row in range(1, 80):
        if val_ws.cell(row=row, column=1).value == "Enterprise value (FCFF perpetuity)":
            assert val_ws.cell(row=row, column=3).data_type == "f"
            break
    else:  # pragma: no cover
        pytest.fail("FCFF perpetuity row not found")

    capm_col = None
    hdr_row: int | None = None
    for row in range(1, 50):
        for col in range(1, 12):
            if tech.cell(row=row, column=col).value == "Expected Ret (CAPM)":
                capm_col = col
                hdr_row = row
                break
        if capm_col is not None:
            break
    assert capm_col is not None and hdr_row is not None
    for row in range(hdr_row + 1, hdr_row + 10):
        if tech.cell(row=row, column=1).value == "AAA":
            assert tech.cell(row=row, column=capm_col).data_type == "f"
            break
    else:  # pragma: no cover
        pytest.fail("FF3 AAA row not found")


def test_valuation_result_model_dump_omits_export_drivers():
    val = _valuation()
    dumped = val.model_dump()
    assert "export_drivers" not in dumped
    assert "per_ticker" in dumped
