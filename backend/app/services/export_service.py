# ExportService — Automated Portfolio Analysis Excel Export
#
# Pedagogy/export only: prices use nominal (unadjusted close) monthly levels when
# available; the S&P 500 sheet uses the ^GSPC index level chain. REST/API analytics
# still follow CONTRACTS adjusted-close + SPY proxy semantics upstream of this builder.
from __future__ import annotations

import io
import json
import time
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas import (
    AnalyticsPerformanceResult,
    ExportRequest,
    OptimizationResult,
    RiskProfile,
    ValuationResult,
)
from app.services.optimize_service import MARKET_PROXY_TICKER


def _scalar_agg_range(col: str, row_start: int, row_end: int) -> str:
    """Plain A1 range for aggregates. Do not wrap in extra parentheses—Excel 365 may
    prepend implicit-intersection (@) to VAR.P, STDEV.P, etc. when the argument is (range).
    """
    return f"{col}{row_start}:{col}{row_end}"


def _excel_scalar(expr: str) -> str:
    """Return plain formula fragments (locale-safe for broader Excel clients)."""
    return expr


def _debug_log(
    run_id: str,
    hypothesis_id: str,
    location: str,
    message: str,
    data: dict[str, Any],
) -> None:
    payload = {
        "sessionId": "102e72",
        "runId": run_id,
        "hypothesisId": hypothesis_id,
        "location": location,
        "message": message,
        "data": data,
        "timestamp": int(time.time() * 1000),
    }
    try:
        log_path = Path(__file__).resolve().parents[3] / "debug-102e72.log"
        with log_path.open("a", encoding="utf-8") as f:
            f.write(json.dumps(payload, ensure_ascii=True) + "\n")
    except OSError:
        pass


def _ref(row: int, col: int = 2) -> str:
    return f"{get_column_letter(col)}{row}"


def _cref(row: int, col: int = 3) -> str:
    return f"{get_column_letter(col)}{row}"


def _dref(row: int, col: int = 4) -> str:
    return f"{get_column_letter(col)}{row}"


def _nominal_px(bar: Any) -> float:
    c = getattr(bar, "close_nominal", None)
    if c is not None:
        return float(c)
    raise ValueError("PriceBar missing close_nominal; cannot write regular close to the workbook")


class ExportService:
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.bold_font = Font(bold=True)
        self.title_font = Font(bold=True, size=14)

    @staticmethod
    def _percent_input_cell(ws, row: int, col: int, val: float | None) -> None:
        c = ws.cell(row=row, column=col, value=val)
        if val is not None:
            c.number_format = "0.00%"

    def build_workbook(
        self,
        optimize: OptimizationResult,
        analytics: AnalyticsPerformanceResult,
        valuation: ValuationResult,
        market_index_bars: list,
        rf_series: list,
        *,
        risk_profile: RiskProfile,
        allow_leverage: bool,
        export_request: ExportRequest,
        aligned_log_returns: pd.DataFrame,
        annualization_factor: int,
        allow_short: bool,
        ff_monthly: pd.DataFrame,
    ) -> io.BytesIO:
        wb = openpyxl.Workbook()
        wb.remove(wb.worksheets[0])

        self._add_fred_graph_sheet(wb, rf_series)
        self._add_market_sheet(wb, market_index_bars)
        self._add_optimize_inputs_sheet(
            wb,
            aligned_log_returns,
            annualization_factor,
            optimize,
            allow_short,
            MARKET_PROXY_TICKER,
        )
        if not ff_monthly.empty:
            self._add_ff3_monthly_sheet(wb, ff_monthly, [s.ticker for s in optimize.stocks])

        tickers = [s.ticker for s in optimize.stocks]
        for ticker in tickers:
            v_block = next((vv for vv in valuation.per_ticker if vv.ticker == ticker), None)
            self._add_ticker_sheet(wb, ticker, v_block, rf_series)

        self._add_chart_sheet(wb, tickers, market_index_bars)
        self._add_regression_sheet(wb, tickers, valuation)
        self._add_valuation_sheet(wb, valuation, export_request, optimize)
        self._add_technical_analysis_sheet(
            wb,
            analytics,
            optimize,
            ff_monthly.empty,
            int(aligned_log_returns.shape[0]),
        )
        self._add_portfolio_org_sheet(
            wb,
            optimize,
            risk_profile,
            allow_leverage,
            allow_short,
            int(aligned_log_returns.shape[0]),
        )

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _add_optimize_inputs_sheet(
        self,
        wb,
        log_rets: pd.DataFrame,
        ann_k: int,
        optimize: OptimizationResult,
        allow_short: bool,
        spy_ticker: str,
    ) -> None:
        """Aligned log returns, tangency weights (Excel 365 LET), ORP series, annual covariance."""
        ws = wb.create_sheet("Optimize Inputs")
        ws["A1"] = "Annualization k"
        ws["B1"] = int(ann_k)
        ws["A2"] = "r_f annual (optimize)"
        rf2 = ws["B2"]
        rf2.value = float(optimize.risk_free_rate)
        rf2.number_format = "0.00%"
        ws["A3"] = "r_f per period"
        ws["B3"] = "=B2/B1"
        ws["B3"].number_format = "0.0000%"
        ws["T1"] = "allow_short"
        ws["U1"] = bool(allow_short)

        tickers = [s.ticker for s in optimize.stocks]
        n_t = len(tickers)
        if not n_t or spy_ticker not in log_rets.columns:
            ws["A5"] = "(insufficient data for aligned returns)"
            return

        idx0 = 2
        first_data_row = 9
        for j, t in enumerate(tickers):
            ws.cell(row=8, column=idx0 + j, value=t).font = self.bold_font
        # Blank separator column so ORP SUMPRODUCT (tickers only) is not adjacent to SPY data.
        spy_col = idx0 + n_t + 1
        orp_col = idx0 + n_t + 2
        ws.cell(row=8, column=spy_col, value=spy_ticker).font = self.bold_font
        ws.cell(row=8, column=orp_col, value="ORP_log").font = self.bold_font
        ws["A8"] = "Date"
        ws["A8"].font = self.bold_font

        last_data_row = first_data_row - 1
        for i, (ts, row) in enumerate(log_rets.iterrows(), first_data_row):
            last_data_row = i
            d = ts.date() if hasattr(ts, "date") else ts
            ws.cell(row=i, column=1, value=d)
            for j, t in enumerate(tickers):
                ws.cell(row=i, column=idx0 + j, value=float(row[t]))
            ws.cell(row=i, column=spy_col, value=float(row[spy_ticker]))
            # idx0 + n_t is intentionally empty (separator column)

        backend_w_row = 55
        for j, t in enumerate(tickers):
            c = get_column_letter(idx0 + j)
            w0 = float(optimize.orp.weights.get(t, 0.0))
            ws.cell(row=backend_w_row, column=idx0 + j, value=w0).number_format = "0.00%"

        for j in range(n_t):
            c = get_column_letter(idx0 + j)
            ws.cell(
                row=6,
                column=idx0 + j,
                value=(
                    f"={_excel_scalar(f'AVERAGE({c}${first_data_row}:{c}${last_data_row})*$B$1')}"
                ),
            )
            ws.cell(row=6, column=idx0 + j).number_format = "0.00%"
        ws.cell(row=6, column=1, value="μ annual (mean log × k)").font = self.bold_font

        spy_l = get_column_letter(spy_col)
        for j in range(n_t):
            c = get_column_letter(idx0 + j)
            ws.cell(
                row=7,
                column=idx0 + j,
                value=(
                    f"=SLOPE({c}${first_data_row}:{c}${last_data_row}-$B$3,"
                    f"{spy_l}${first_data_row}:{spy_l}${last_data_row}-$B$3)"
                ),
            )
        ws.cell(row=7, column=1, value="β vs SPY (excess on excess)").font = self.bold_font

        cov_r0 = last_data_row + 3
        start_cov = orp_col + 2
        exc_col = start_cov + n_t  # first column to the right of the n x n cov block
        ws.cell(
            row=cov_r0 - 1, column=1, value="Covariance annual (COVARIANCE.S × k)"
        ).font = self.bold_font
        for ii in range(n_t):
            ci = get_column_letter(idx0 + ii)
            for jj in range(n_t):
                cj = get_column_letter(idx0 + jj)
                cell = ws.cell(
                    row=cov_r0 + ii,
                    column=start_cov + jj,
                    value=(
                        f"={_excel_scalar(f'COVARIANCE.S(${ci}${first_data_row}:${ci}${last_data_row},${cj}${first_data_row}:${cj}${last_data_row})*$B$1')}"
                    ),
                )
                cell.number_format = "0.0000"
        for ii in range(n_t):
            cmu = get_column_letter(idx0 + ii)
            ws.cell(
                row=cov_r0 + ii,
                column=exc_col,
                value=f"={cmu}$6-$B$2",
            )

        ws.cell(row=5, column=1, value="w_ORP tangency / backend").font = self.bold_font

        for j, t in enumerate(tickers):
            c = get_column_letter(idx0 + j)
            w_backend = float(optimize.orp.weights.get(t, 0.0))
            cell_w = ws.cell(row=5, column=idx0 + j, value=w_backend)
            cell_w.number_format = "0.00%"

        tl = get_column_letter(idx0)
        t_last = get_column_letter(idx0 + n_t - 1)
        for r in range(first_data_row, last_data_row + 1):
            orp_formula = (
                f"={_excel_scalar(f'SUMPRODUCT(${tl}$5:${t_last}$5,${tl}{r}:${t_last}{r})')}"
            )
            ws.cell(
                row=r,
                column=orp_col,
                value=orp_formula,
            )

    def _add_ff3_monthly_sheet(self, wb, ff: pd.DataFrame, tickers: list[str]) -> None:
        ws = wb.create_sheet("FF3 Monthly")
        headers = ["Date", "RF", "Mkt_RF", "SMB", "HML", *tickers]
        for j, h in enumerate(headers, 1):
            ws.cell(row=1, column=j, value=h).font = self.bold_font
        for i, (ts, row) in enumerate(ff.iterrows(), 2):
            d = ts.date() if hasattr(ts, "date") else ts
            ws.cell(row=i, column=1, value=d)
            ws.cell(row=i, column=2, value=float(row["RF"]))
            ws.cell(row=i, column=3, value=float(row["Mkt_RF"]))
            ws.cell(row=i, column=4, value=float(row["SMB"]))
            ws.cell(row=i, column=5, value=float(row["HML"]))
            for j, t in enumerate(tickers, 6):
                ws.cell(row=i, column=j, value=float(row[t]))

    def _add_fred_graph_sheet(self, wb, rf_series):
        ws = wb.create_sheet("Fred Graph")
        headers = ["observation_date", "DGS3MO", "rf"]
        for i, h in enumerate(headers, 1):
            cell = ws.cell(row=5, column=i, value=h)
            cell.font = self.bold_font
            cell.border = self.border

        sorted_rf = sorted(rf_series, key=lambda x: x["date"])
        for i, item in enumerate(sorted_rf, 6):
            ws.cell(row=i, column=1, value=item["date"])
            ws.cell(row=i, column=2, value=item["rate"] * 100)
            ws.cell(row=i, column=3, value=f"=B{i}/(100*12)")

        ws["F6"] = "Average rf"
        ws["G6"] = f"=AVERAGE(C6:C{5 + len(sorted_rf)})"

    def _add_market_sheet(self, wb, market_index_bars):
        ws = wb.create_sheet("S&P 500")
        headers = ["Date", "P_GSPC (regular close)", "P_prev", "RM"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h).font = self.bold_font

        sorted_bars = sorted(market_index_bars, key=lambda x: x.date)
        for i, bar in enumerate(sorted_bars, 2):
            ws.cell(row=i, column=1, value=bar.date)
            ws.cell(row=i, column=2, value=_nominal_px(bar))
            if i > 2:
                ws.cell(row=i, column=3, value=f"=B{i - 1}")
                ws.cell(row=i, column=4, value=f"=(B{i}-C{i})/C{i}")

    def _add_ticker_sheet(self, wb, ticker, v_block, rf_series):
        ws = wb.create_sheet(ticker)
        headers = [
            "Date",
            f"P_{ticker} (regular close)",
            "P_prev",
            f"R_{ticker}",
            "",
            "RF",
            "RM",
            "",
            f"1+R_{ticker}",
            "Risk Premium",
        ]
        for i, h in enumerate(headers, 1):
            if h:
                ws.cell(row=1, column=i, value=h).font = self.bold_font

        if not v_block or not v_block.historical_prices:
            return

        prices = sorted(v_block.historical_prices, key=lambda x: x.date)
        for i, p in enumerate(prices, 2):
            ws.cell(row=i, column=1, value=p.date)
            ws.cell(row=i, column=2, value=_nominal_px(p))
            if i > 2:
                ws.cell(row=i, column=3, value=f"=B{i - 1}")
                ws.cell(row=i, column=4, value=f"=(B{i}-C{i})/C{i}")

            ws.cell(row=i, column=6, value=f"='Fred Graph'!C{i + 4}")
            ws.cell(row=i, column=7, value=f"='S&P 500'!D{i}")
            if i > 2:
                ws.cell(row=i, column=9, value=f"=1+D{i}")
                ws.cell(row=i, column=10, value=f"=D{i}-F{i}")

        last_row = len(prices) + 1
        ws["L1"] = "Metric"
        ws["M1"] = "10 Years"
        ws["N1"] = "5 Years"
        ws["O1"] = "3 Years"
        for j in range(12, 16):
            ws.cell(row=1, column=j).font = self.bold_font

        metrics = [
            ("Arithmetic Monthly Return", "D"),
            ("Geometric Monthly Return", "I"),
            ("Variance", "D"),
            ("Standard Deviation", "D"),
            ("Monthly Risk Premium", "J"),
        ]

        for i, (m_name, col) in enumerate(metrics, 2):
            ws.cell(row=i, column=12, value=m_name)
            rng_full = _scalar_agg_range(col, 3, last_row)
            if m_name == "Geometric Monthly Return":
                ws.cell(row=i, column=13, value=f"={_excel_scalar(f'GEOMEAN({rng_full})')}-1")
            elif m_name == "Standard Deviation":
                ws.cell(row=i, column=13, value=f"={_excel_scalar(f'STDEVP({rng_full})')}")
            else:
                inner = (
                    f"AVERAGE({rng_full})"
                    if "Return" in m_name or "Premium" in m_name
                    else f"VARP({rng_full})"
                )
                ws.cell(
                    row=i,
                    column=13,
                    value=f"={_excel_scalar(inner)}",
                )

            r5 = max(3, last_row - 59)
            rng_5 = _scalar_agg_range(col, r5, last_row)
            if m_name == "Geometric Monthly Return":
                ws.cell(row=i, column=14, value=f"={_excel_scalar(f'GEOMEAN({rng_5})')}-1")
            elif m_name == "Standard Deviation":
                ws.cell(row=i, column=14, value=f"={_excel_scalar(f'STDEVP({rng_5})')}")
            else:
                inner5 = (
                    f"AVERAGE({rng_5})"
                    if "Return" in m_name or "Premium" in m_name
                    else f"VARP({rng_5})"
                )
                ws.cell(row=i, column=14, value=f"={_excel_scalar(inner5)}")

            r3 = max(3, last_row - 35)
            rng_3 = _scalar_agg_range(col, r3, last_row)
            if m_name == "Geometric Monthly Return":
                ws.cell(row=i, column=15, value=f"={_excel_scalar(f'GEOMEAN({rng_3})')}-1")
            elif m_name == "Standard Deviation":
                ws.cell(row=i, column=15, value=f"={_excel_scalar(f'STDEVP({rng_3})')}")
            else:
                inner3 = (
                    f"AVERAGE({rng_3})"
                    if "Return" in m_name or "Premium" in m_name
                    else f"VARP({rng_3})"
                )
                ws.cell(row=i, column=15, value=f"={_excel_scalar(inner3)}")

    def _add_chart_sheet(self, wb, tickers, market_index_bars):
        ws = wb.create_sheet("Chart")
        ws.cell(row=1, column=1, value="Date").font = self.bold_font
        for j, t in enumerate(tickers, 2):
            ws.cell(row=1, column=j, value=t).font = self.bold_font
        spy_col_label = len(tickers) + 2
        rf_col_label = len(tickers) + 3
        ws.cell(row=1, column=spy_col_label, value="S&P 500 Index").font = self.bold_font
        ws.cell(row=1, column=rf_col_label, value="RF").font = self.bold_font

        for j in range(2, len(tickers) + 5):
            ws.cell(row=2, column=j, value=100)

        dates = [b.date for b in sorted(market_index_bars, key=lambda x: x.date)]
        for i, d in enumerate(dates, 2):
            ws.cell(row=i, column=1, value=d)
            if i > 2:
                for j, ticker_name in enumerate(tickers, 2):
                    ws.cell(
                        row=i,
                        column=j,
                        value=f"={get_column_letter(j)}{i - 1}*(1+'{ticker_name}'!D{i})",
                    )
                ws.cell(
                    row=i,
                    column=spy_col_label,
                    value=f"={get_column_letter(spy_col_label)}{i - 1}*(1+'S&P 500'!D{i})",
                )
                ws.cell(
                    row=i,
                    column=rf_col_label,
                    value=f"={get_column_letter(rf_col_label)}{i - 1}*(1+'Fred Graph'!C{i + 4})",
                )

    def _add_regression_sheet(self, wb, tickers, valuation):
        ws = wb.create_sheet("1 Regression")
        col_off = 1

        for ticker in tickers:
            block = next((vv for vv in valuation.per_ticker if vv.ticker == ticker), None)
            if not block or not block.historical_prices:
                col_off += 5
                continue
            last_row = len(block.historical_prices) + 1

            ws.cell(
                row=1, column=col_off, value=f"{ticker} Regression Output"
            ).font = self.title_font
            ws.cell(row=3, column=col_off, value="Regression Statistics").font = self.bold_font

            y = f"'{ticker}'!J3:J{last_row}"
            x = f"'{ticker}'!G3:G{last_row}"

            stats = [
                ("Multiple R", f"=SQRT(RSQ({y},{x}))"),
                ("R Square", f"=RSQ({y},{x})"),
                ("Adjusted R Square", f"=1-(1-RSQ({y},{x}))*(COUNT({y})-1)/(COUNT({y})-2)"),
                ("Standard Error", f"=STEYX({y},{x})"),
                ("Observations", f"=COUNT({y})"),
            ]
            for i, (lab, form) in enumerate(stats, 4):
                ws.cell(row=i, column=col_off, value=lab)
                ws.cell(row=i, column=col_off + 1, value=form)

            ws.cell(row=16, column=col_off + 1, value="Coefficients").font = self.bold_font
            ws.cell(row=17, column=col_off, value="Intercept")
            ws.cell(row=17, column=col_off + 1, value=f"=INTERCEPT({y},{x})")
            ws.cell(row=18, column=col_off, value="Market (Beta)")
            ws.cell(row=18, column=col_off + 1, value=f"=SLOPE({y},{x})")

            col_off += 5

    def _add_valuation_sheet(
        self,
        wb,
        valuation: ValuationResult,
        req: ExportRequest,
        optimize: OptimizationResult,
    ) -> None:
        ws = wb.create_sheet("Valuation and Earnings Analysis")

        ws.cell(row=1, column=3, value="Modeled (Excel formula)").font = self.bold_font
        ws.cell(row=1, column=4, value="Backend output (comparison)").font = self.bold_font
        ws.cell(row=2, column=1, value="Field").font = self.bold_font
        ws.cell(row=2, column=2, value="Input (API / export)")
        ws.cell(row=2, column=3, value="Modeled formula")
        ws.cell(row=2, column=4, value="Backend snapshot")

        r = 3
        drivers_by = {x.ticker: x for x in valuation.export_drivers}
        modeled_ke_rows: dict[str, int] = {}
        modeled_wacc_rows: dict[str, int] = {}
        modeled_fcff_rows: dict[str, int] = {}
        modeled_fcfe_rows: dict[str, int] = {}
        for v in valuation.per_ticker:
            d = drivers_by.get(v.ticker)
            ws.cell(
                row=r, column=1, value=f"{v.ticker} Fundamental Valuation"
            ).font = self.title_font
            r += 1

            ws.cell(
                row=r, column=1, value="Export overrides (annualized decimals)"
            ).font = self.bold_font
            r += 1

            r_ov_wacc = r
            ws.cell(row=r, column=1, value="WACC Override")
            self._percent_input_cell(ws, r, 2, req.wacc)
            r += 1
            r_ov_gfcfe = r
            ws.cell(row=r, column=1, value="FCFE growth Override")
            self._percent_input_cell(ws, r, 2, req.fcff_growth)
            r += 1
            r_ov_gt = r
            ws.cell(row=r, column=1, value="FCFF terminal growth Override")
            self._percent_input_cell(ws, r, 2, req.fcff_terminal_growth)
            r += 1
            r_ov_gdd = r
            ws.cell(row=r, column=1, value="Gordon dividend growth Override")
            self._percent_input_cell(ws, r, 2, req.ddm_gordon_g)
            r += 1
            r_ov_ke = r
            ws.cell(row=r, column=1, value="k_e Override")
            self._percent_input_cell(ws, r, 2, req.cost_of_equity_override)
            r += 2

            ws.cell(
                row=r, column=1, value="Fundamentals & DCF drivers (inputs col B; modeled col C)"
            ).font = self.bold_font
            r += 1

            if d is not None:
                r_rf_capm = r
                ws.cell(row=r, column=1, value="Risk-free r_f (annual, CAPM)")
                ws.cell(row=r, column=2, value=float(d.risk_free_annual)).number_format = "0.00%"
                r += 1
                r_beta = r
                ws.cell(row=r, column=1, value="Beta (CAPM)")
                if d.beta is not None:
                    ws.cell(row=r, column=2, value=float(d.beta)).number_format = "0.000"
                r += 1
                r_mrp = r
                ws.cell(row=r, column=1, value="Market risk premium (assumption)")
                if d.market_risk_premium is not None:
                    ws.cell(row=r, column=2, value=float(d.market_risk_premium)).number_format = "0.00%"
                r += 1
                r_fun = r
                ws.cell(row=r, column=1, value="Financial sector (FCFF/FCFE omitted if TRUE)")
                ws.cell(row=r, column=2, value="TRUE" if d.financial_unsafe else "FALSE")
                r += 1
                r_ebit_b = r
                ws.cell(row=r, column=1, value="EBIT")
                if d.ebit is not None:
                    ws.cell(row=r, column=2, value=float(d.ebit)).number_format = "#,##0"
                r += 1
                r_taxb = r
                ws.cell(row=r, column=1, value="Effective tax rate Tc")
                if d.tax_rate is not None:
                    ws.cell(row=r, column=2, value=float(d.tax_rate)).number_format = "0.00%"
                r += 1
                r_deprb = r
                ws.cell(row=r, column=1, value="D&A / depreciation")
                if d.depreciation is not None:
                    ws.cell(row=r, column=2, value=float(d.depreciation)).number_format = "#,##0"
                r += 1
                r_capxb = r
                ws.cell(row=r, column=1, value="CapEx (positive magnitude)")
                if d.capex is not None:
                    ws.cell(row=r, column=2, value=float(d.capex)).number_format = "#,##0"
                r += 1
                r_dnwc = r
                ws.cell(row=r, column=1, value="ΔNWC (current − prior)")
                if d.delta_nwc is not None:
                    ws.cell(row=r, column=2, value=float(d.delta_nwc)).number_format = "#,##0"
                r += 1
                r_intb = r
                ws.cell(row=r, column=1, value="Interest expense (positive magnitude)")
                if d.interest_expense is not None:
                    ws.cell(row=r, column=2, value=float(d.interest_expense)).number_format = "#,##0"
                r += 1
                r_nbb = r
                ws.cell(row=r, column=1, value="Net borrowing (Δ debt)")
                if d.net_borrowing is not None:
                    ws.cell(row=r, column=2, value=float(d.net_borrowing)).number_format = "#,##0"
                r += 1
                r_mcb = r
                ws.cell(row=r, column=1, value="Market cap E")
                if d.market_cap is not None:
                    ws.cell(row=r, column=2, value=float(d.market_cap)).number_format = "#,##0"
                r += 1
                r_debtb = r
                ws.cell(row=r, column=1, value="Interest-bearing debt D")
                if d.total_debt is not None:
                    ws.cell(row=r, column=2, value=float(d.total_debt)).number_format = "#,##0"
                r += 1
                r_kdb = r
                ws.cell(row=r, column=1, value="Cost of debt kd (pretax)")
                if d.cost_of_debt_pretax is not None:
                    ws.cell(
                        row=r, column=2, value=float(d.cost_of_debt_pretax)
                    ).number_format = "0.00%"
                r += 1
            else:
                r_rf_capm = r
                ws.cell(row=r, column=1, value="(No export drivers for ticker)")
                r += 1
                stub = r
                r_beta = stub
                r_mrp = stub
                r_fun = stub
                r_ebit_b = stub
                r_taxb = stub
                r_deprb = stub
                r_capxb = stub
                r_dnwc = stub
                r_intb = stub
                r_nbb = stub
                r_mcb = stub
                r_debtb = stub
                r_kdb = stub

            r_ke = r
            ws.cell(row=r, column=1, value="k_e (modeled CAPM)")
            if d is not None:
                ws.cell(
                    row=r,
                    column=3,
                    value=(
                        f"=IF(NOT(ISBLANK({_ref(r_ov_ke)})),{_ref(r_ov_ke)},"
                        f"IF(AND(ISNUMBER({_ref(r_rf_capm)}),ISNUMBER({_ref(r_beta)}),ISNUMBER({_ref(r_mrp)})),"
                        f"{_ref(r_rf_capm)}+{_ref(r_beta)}*{_ref(r_mrp)},\"\"))"
                    ),
                )
                ws.cell(row=r, column=3).number_format = "0.00%"
            elif v.cost_of_equity is not None:
                ws.cell(row=r, column=3, value=float(v.cost_of_equity)).number_format = "0.00%"
            self._percent_input_cell(
                ws,
                r,
                4,
                float(v.cost_of_equity) if v.cost_of_equity is not None else None,
            )
            modeled_ke_rows[v.ticker] = r
            r += 1
            r_wacc = r
            ws.cell(row=r, column=1, value="WACC (modeled)")
            if d is not None:
                ev_plus_d = f"({_ref(r_mcb)}+{_ref(r_debtb)})"
                ws.cell(
                    row=r,
                    column=3,
                    value=(
                        f"=IF(NOT(ISBLANK({_ref(r_ov_wacc)})),{_ref(r_ov_wacc)},"
                        f"IF(AND(ISNUMBER({_ref(r_mcb)}),ISNUMBER({_ref(r_debtb)}),{ev_plus_d}>0,"
                        f"ISNUMBER({_cref(r_ke)}),ISNUMBER({_ref(r_kdb)}),ISNUMBER({_ref(r_taxb)})),"
                        f"{_ref(r_mcb)}/{ev_plus_d}*{_cref(r_ke)}+"
                        f"{_ref(r_debtb)}/{ev_plus_d}*{_ref(r_kdb)}*(1-{_ref(r_taxb)}),"
                        f"IF(AND(ISNUMBER({_ref(r_mcb)}),{_ref(r_mcb)}>0,{_ref(r_debtb)}=0,ISNUMBER({_cref(r_ke)})),"
                        f"{_cref(r_ke)},\"\")))"
                    ),
                )
                ws.cell(row=r, column=3).number_format = "0.00%"
            if v.wacc is not None:
                ws.cell(row=r, column=4, value=float(v.wacc)).number_format = "0.00%"
            modeled_wacc_rows[v.ticker] = r_wacc
            r += 1
            r_fcff = r
            ws.cell(row=r, column=1, value="FCFF (modeled from EBIT)")
            if d is not None:
                ws.cell(
                    row=r,
                    column=3,
                    value=(
                        f'=IF(OR({_ref(r_fun)}="TRUE",NOT(ISNUMBER({_ref(r_ebit_b)})),'
                        f"NOT(ISNUMBER({_ref(r_taxb)})),NOT(ISNUMBER({_ref(r_deprb)})),"
                        f'NOT(ISNUMBER({_ref(r_capxb)})),NOT(ISNUMBER({_ref(r_dnwc)}))),"",'
                        f"{_ref(r_ebit_b)}*(1-{_ref(r_taxb)})+{_ref(r_deprb)}"
                        f"-{_ref(r_capxb)}-{_ref(r_dnwc)})"
                    ),
                )
                ws.cell(row=r, column=3).number_format = "#,##0"
            if v.fcff is not None:
                ws.cell(row=r, column=4, value=float(v.fcff)).number_format = "#,##0"
            modeled_fcff_rows[v.ticker] = r_fcff
            r += 1
            r_fcfe = r
            ws.cell(row=r, column=1, value="FCFE (modeled from FCFF)")
            if d is not None:
                ws.cell(
                    row=r,
                    column=3,
                    value=(
                        f'=IF(OR(NOT(ISNUMBER({_cref(r_fcff)})),NOT(ISNUMBER({_ref(r_intb)})),'
                        f'NOT(ISNUMBER({_ref(r_taxb)})),NOT(ISNUMBER({_ref(r_nbb)}))),"",'
                        f"{_cref(r_fcff)}-{_ref(r_intb)}*(1-{_ref(r_taxb)})+{_ref(r_nbb)})"
                    ),
                )
                ws.cell(row=r, column=3).number_format = "#,##0"
            if v.fcfe is not None:
                ws.cell(row=r, column=4, value=float(v.fcfe)).number_format = "#,##0"
            modeled_fcfe_rows[v.ticker] = r_fcfe
            r += 1
            r_gsus = r
            ws.cell(row=r, column=1, value="Sustainable growth g")
            sg = ws.cell(row=r, column=2)
            if v.sustainable_growth_rate is not None:
                sg.number_format = "0.00%"
                sg.value = float(v.sustainable_growth_rate)
            r += 2

            ws.cell(row=r, column=1, value="DCF & growth intermediates").font = self.bold_font
            r += 1

            r_fcfe_g = r
            ws.cell(row=r, column=1, value="FCFE perpetual g (Excel)")
            c_fcfg = ws.cell(row=r, column=3)
            c_fcfg.value = (
                f"=IF(NOT(ISBLANK({_ref(r_ov_gfcfe)})),{_ref(r_ov_gfcfe)},"
                f'IF(ISBLANK({_ref(r_gsus)}),"",{_ref(r_gsus)}))'
            )
            c_fcfg.number_format = "0.00%"
            r += 1

            r_term = r
            ws.cell(row=r, column=1, value="FCFF terminal g (Excel)")
            c_ter = ws.cell(row=r, column=3)
            c_ter.value = f'=IF(ISBLANK({_ref(r_ov_gt)}),"",{_ref(r_ov_gt)})'
            c_ter.number_format = "0.00%"
            r += 1

            ws.cell(row=r, column=1, value="Gordon dividend g (Excel)")
            c_gordon = ws.cell(row=r, column=3)
            c_gordon.value = (
                f"=IF(NOT(ISBLANK({_ref(r_ov_gdd)})),{_ref(r_ov_gdd)},"
                f'IF(ISBLANK({_ref(r_gsus)}),"",{_ref(r_gsus)}))'
            )
            c_gordon.number_format = "0.00%"
            r += 1

            r_ke_use = r
            ws.cell(row=r, column=1, value="k_e used (override or modeled)")
            c_ke_use = ws.cell(row=r, column=3)
            c_ke_use.value = f"=IF(NOT(ISBLANK({_ref(r_ov_ke)})),{_ref(r_ov_ke)},{_cref(r_ke)})"
            c_ke_use.number_format = "0.00%"
            r += 1

            r_w_use = r
            ws.cell(row=r, column=1, value="WACC used for FCFF PV (override or modeled)")
            c_w_use = ws.cell(row=r, column=3)
            c_w_use.value = (
                f"=IF(NOT(ISBLANK({_ref(r_ov_wacc)})),{_ref(r_ov_wacc)},{_cref(r_wacc)})"
            )
            c_w_use.number_format = "0.00%"
            r += 1

            ws.cell(row=r, column=1, value="Enterprise value (FCFF perpetuity)")
            c_ef = ws.cell(row=r, column=3)
            c_ef.value = (
                f"=IF(AND(ISNUMBER({_cref(r_fcff)}),{_cref(r_w_use)}>{_cref(r_term)}),"
                f"{_cref(r_fcff)}*(1+{_cref(r_term)})"
                f'/({_cref(r_w_use)}-{_cref(r_term)}),"")'
            )
            c_ef.number_format = "#,##0"
            r += 1

            ws.cell(row=r, column=1, value="FCFF g used (terminal)")
            c_fcff_g_used = ws.cell(row=r, column=3, value=f"={_cref(r_term)}")
            c_fcff_g_used.number_format = "0.00%"
            if req.fcff_terminal_growth is not None:
                ws.cell(row=r, column=4, value=float(req.fcff_terminal_growth)).number_format = "0.00%"
            r += 1

            ws.cell(row=r, column=1, value="FCFF WACC used")
            c_fcff_wacc_used = ws.cell(row=r, column=3, value=f"={_cref(r_w_use)}")
            c_fcff_wacc_used.number_format = "0.00%"
            if v.wacc is not None:
                ws.cell(row=r, column=4, value=float(v.wacc)).number_format = "0.00%"
            r += 1

            ws.cell(row=r, column=1, value="Equity value from FCFE (perpetuity)")
            c_ee = ws.cell(row=r, column=3)
            c_ee.value = (
                f"=IF(AND(ISNUMBER({_cref(r_fcfe)}),{_cref(r_ke_use)}>{_cref(r_fcfe_g)}),"
                f"{_cref(r_fcfe)}*(1+{_cref(r_fcfe_g)})"
                f'/({_cref(r_ke_use)}-{_cref(r_fcfe_g)}),"")'
            )
            c_ee.number_format = "#,##0"
            r += 1

            ws.cell(
                row=r, column=1, value="Multiples & intrinsic value (API snapshot in col D)"
            ).font = self.bold_font
            r += 1

            ws.cell(row=r, column=1, value="P/E (trailing)")
            pe_d = ws.cell(row=r, column=4)
            if v.price_to_earnings is not None:
                pe_d.value = float(v.price_to_earnings)
                pe_d.number_format = "0.0"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="P/B")
            pb_d = ws.cell(row=r, column=4)
            if v.price_to_book is not None:
                pb_d.value = float(v.price_to_book)
                pb_d.number_format = "0.0"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="ROE")
            roe_d = ws.cell(row=r, column=4)
            if v.roe is not None:
                roe_d.value = float(v.roe)
                roe_d.number_format = "0.00%"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="Gross margin")
            gm_d = ws.cell(row=r, column=4)
            if v.gross_margin is not None:
                gm_d.value = float(v.gross_margin)
                gm_d.number_format = "0.00%"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="Operating margin")
            om_d = ws.cell(row=r, column=4)
            if v.operating_margin is not None:
                om_d.value = float(v.operating_margin)
                om_d.number_format = "0.00%"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="FCFF value per share")
            vps_f = ws.cell(row=r, column=4)
            if v.fcff_value_per_share is not None:
                vps_f.value = float(v.fcff_value_per_share)
                vps_f.number_format = "$#,##0.00"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="FCFE value per share")
            vps_e = ws.cell(row=r, column=4)
            if v.fcfe_value_per_share is not None:
                vps_e.value = float(v.fcfe_value_per_share)
                vps_e.number_format = "$#,##0.00"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="Gordon DDM value per share")
            ddm1_d = ws.cell(row=r, column=4)
            if v.ddm_gordon is not None:
                ddm1_d.value = float(v.ddm_gordon)
                ddm1_d.number_format = "$#,##0.00"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 1

            ws.cell(row=r, column=1, value="2-stage DDM value per share")
            ddm2_d = ws.cell(row=r, column=4)
            if v.ddm_two_stage is not None:
                ddm2_d.value = float(v.ddm_two_stage)
                ddm2_d.number_format = "$#,##0.00"
            ws.cell(row=r, column=3, value=f'=IF(ISNUMBER({_dref(r)}),{_dref(r)},"")')
            r += 2

        if modeled_ke_rows:
            r += 2
            ws.cell(
                row=r, column=1, value="Portfolio — ORP-weighted modeled rows"
            ).font = self.title_font
            r += 1
            tickers_o = [s.ticker for s in optimize.stocks]

            def _wsum(row_map: dict[str, int]) -> str | None:
                parts: list[str] = []
                for i, t in enumerate(tickers_o):
                    rr = row_map.get(t)
                    if rr is None:
                        continue
                    letter = get_column_letter(2 + i)
                    parts.append(f"'Optimize Inputs'!{letter}$5*C{rr}")
                if not parts:
                    return None
                return "=" + "+".join(parts)

            for label, rmap, fmt in (
                ("k_e ORP-weighted", modeled_ke_rows, "0.00%"),
                ("WACC ORP-weighted", modeled_wacc_rows, "0.00%"),
                ("FCFF ORP-weighted", modeled_fcff_rows, "#,##0"),
                ("FCFE ORP-weighted", modeled_fcfe_rows, "#,##0"),
            ):
                fsum = _wsum(rmap)
                if fsum:
                    ws.cell(row=r, column=1, value=label)
                    c = ws.cell(row=r, column=3, value=fsum)
                    c.number_format = fmt
                    r += 1

    def _add_technical_analysis_sheet(
        self,
        wb,
        analytics: AnalyticsPerformanceResult,
        optimize: OptimizationResult,
        ff3_missing: bool,
        n_obs: int,
    ) -> None:
        ws = wb.create_sheet("Technical Analysis")
        ws.cell(row=1, column=1, value="Technical Performance Metrics").font = self.title_font
        ws.cell(row=2, column=1, value="Field").font = self.bold_font
        ws.cell(row=2, column=2, value="Backend (comparison)").font = self.bold_font
        ws.cell(row=2, column=3, value="Excel formula (Optimize Inputs)").font = self.bold_font

        n_t = len(optimize.stocks)
        first_data_row = 9
        last_data_row = first_data_row + n_obs - 1
        orp_c = get_column_letter(2 + n_t + 2)
        spy_c = get_column_letter(2 + n_t + 1)
        oi = "Optimize Inputs"

        r = 3
        ws.cell(row=r, column=1, value="ORP Portfolio Performance").font = self.bold_font
        r += 1

        ws.cell(row=r, column=1, value="Jensen's Alpha (Annual)")
        ws.cell(row=r, column=2, value=float(analytics.orp.jensen_alpha)).number_format = "0.00%"
        jensen_inner = (
            f"INTERCEPT('{oi}'!{orp_c}{first_data_row}:{orp_c}{last_data_row}-'{oi}'!$B$3,"
            f"'{oi}'!{spy_c}{first_data_row}:{spy_c}{last_data_row}-'{oi}'!$B$3)*'{oi}'!$B$1"
        )
        ws.cell(row=r, column=3, value=f"={_excel_scalar(jensen_inner)}")
        ws.cell(row=r, column=3).number_format = "0.00%"
        # #region agent log
        _debug_log(
            "debug-pre-fix",
            "H3",
            "export_service.py:_add_technical_analysis_sheet",
            "jensen_formula_and_backend",
            {
                "cell": f"C{r}",
                "formula": ws.cell(row=r, column=3).value,
                "backend": float(analytics.orp.jensen_alpha),
                "nObs": int(n_obs),
            },
        )
        # #endregion
        r += 1

        ws.cell(row=r, column=1, value="Treynor Ratio")
        ws.cell(row=r, column=2, value=float(analytics.orp.treynor)).number_format = "0.000"
        treynor_avg = "'Portfolio Organization'!B8"
        treynor_betas = _excel_scalar(
            f"SUMPRODUCT('{oi}'!{get_column_letter(2)}$5:"
            f"'{oi}'!{get_column_letter(2 + n_t - 1)}$5,"
            f"'{oi}'!{get_column_letter(2)}$7:'{oi}'!{get_column_letter(2 + n_t - 1)}$7)"
        )
        ws.cell(
            row=r,
            column=3,
            value=(f"=({treynor_avg}*'{oi}'!$B$1-'{oi}'!$B$2)/{treynor_betas}"),
        )
        ws.cell(row=r, column=3).number_format = "0.000"
        # #region agent log
        _debug_log(
            "debug-pre-fix",
            "H2",
            "export_service.py:_add_technical_analysis_sheet",
            "treynor_formula_and_backend",
            {
                "cell": f"C{r}",
                "formula": ws.cell(row=r, column=3).value,
                "backend": float(analytics.orp.treynor),
                "nObs": int(n_obs),
            },
        )
        # #endregion
        r += 1

        ws.cell(row=r, column=1, value="Sharpe Ratio (ORP)")
        sh_backend = float(optimize.orp.sharpe)
        ws.cell(row=r, column=2, value=sh_backend).number_format = "0.000"
        sharpe_avg = "'Portfolio Organization'!B8"
        sharpe_sd = "'Portfolio Organization'!B10"
        ws.cell(
            row=r,
            column=3,
            value=f"=({sharpe_avg}-'{oi}'!$B$2)/{sharpe_sd}",
        )
        ws.cell(row=r, column=3).number_format = "0.000"
        # #region agent log
        _debug_log(
            "debug-pre-fix",
            "H1",
            "export_service.py:_add_technical_analysis_sheet",
            "sharpe_formula_and_backend",
            {
                "cell": f"C{r}",
                "formula": ws.cell(row=r, column=3).value,
                "backend": float(optimize.orp.sharpe),
                "sharpeAvgRef": sharpe_avg,
                "sharpeSdRef": sharpe_sd,
            },
        )
        # #endregion
        r += 1

        r += 1
        ws.cell(
            row=r, column=1, value="Fama-French 3-Factor Model (FF3 Monthly sheet)"
        ).font = self.bold_font
        r += 1
        headers = [
            "Ticker",
            "Beta Mkt",
            "Beta SMB",
            "Beta HML",
            "Alpha (Annual)",
            "Expected Ret (FF3)",
            "Expected Ret (CAPM)",
            "Backend FF3",
            "Backend CAPM",
        ]
        for j, h in enumerate(headers, 1):
            ws.cell(row=r, column=j, value=h).font = self.bold_font
        r += 1

        if ff3_missing:
            ws.cell(
                row=r, column=1, value="(No aligned monthly history / factors for FF3 in workbook)"
            )
            return

        ff_ws = wb["FF3 Monthly"]
        ff_last_row = ff_ws.max_row
        if ff_last_row < 2:
            ws.cell(row=r, column=1, value="(FF3 Monthly empty)")
            return

        fr, lr = 2, ff_last_row
        ff_row = r
        for f in analytics.fama_french:
            t = f.ticker
            hdr = [ff_ws.cell(row=1, column=j).value for j in range(1, ff_ws.max_column + 1)]
            try:
                tc = hdr.index(t) + 1
            except ValueError:
                continue
            tc_letter = get_column_letter(tc)
            y_rng = f"'FF3 Monthly'!{tc_letter}{fr}:{tc_letter}{lr}-'FF3 Monthly'!B{fr}:B{lr}"
            x_rng = f"'FF3 Monthly'!C{fr}:E{lr}"
            ws.cell(row=ff_row, column=1, value=t)
            ws.cell(
                row=ff_row,
                column=2,
                value=f"=INDEX(LINEST({y_rng},{x_rng},TRUE,TRUE),1,3)",
            )
            ws.cell(
                row=ff_row,
                column=3,
                value=f"=INDEX(LINEST({y_rng},{x_rng},TRUE,TRUE),1,2)",
            )
            ws.cell(
                row=ff_row,
                column=4,
                value=f"=INDEX(LINEST({y_rng},{x_rng},TRUE,TRUE),1,1)",
            )
            ws.cell(
                row=ff_row,
                column=5,
                value=f"=INDEX(LINEST({y_rng},{x_rng},TRUE,TRUE),1,4)*12",
            )
            ws.cell(row=ff_row, column=5).number_format = "0.00%"
            ws.cell(
                row=ff_row,
                column=6,
                value=(
                    f"=12*(AVERAGE('FF3 Monthly'!B{fr}:B{lr})+"
                    f"B{ff_row}*AVERAGE('FF3 Monthly'!C{fr}:C{lr})+"
                    f"C{ff_row}*AVERAGE('FF3 Monthly'!D{fr}:D{lr})+"
                    f"D{ff_row}*AVERAGE('FF3 Monthly'!E{fr}:E{lr}))"
                ),
            )
            ws.cell(row=ff_row, column=6).number_format = "0.00%"
            ws.cell(
                row=ff_row,
                column=7,
                value=(
                    f"=12*(AVERAGE('FF3 Monthly'!B{fr}:B{lr})+"
                    f"B{ff_row}*AVERAGE('FF3 Monthly'!C{fr}:C{lr}))"
                ),
            )
            ws.cell(row=ff_row, column=7).number_format = "0.00%"
            ws.cell(
                row=ff_row, column=8, value=float(f.expected_return_ff3)
            ).number_format = "0.00%"
            ws.cell(
                row=ff_row, column=9, value=float(f.expected_return_capm)
            ).number_format = "0.00%"
            ff_row += 1

    def _add_portfolio_org_sheet(
        self,
        wb,
        optimize: OptimizationResult,
        risk_profile: RiskProfile,
        allow_leverage: bool,
        allow_short: bool,
        n_obs: int,
    ) -> None:
        ws = wb.create_sheet("Portfolio Organization")
        n_t = len(optimize.stocks)
        first_data_row = 9
        last_data_row = first_data_row + n_obs - 1
        orp_col = 2 + n_t + 2
        start_cov = orp_col + 2
        cov_r0 = last_data_row + 3
        cov_tl = f"'Optimize Inputs'!${get_column_letter(start_cov)}${cov_r0}"
        cov_br = f"'Optimize Inputs'!${get_column_letter(start_cov + n_t - 1)}${cov_r0 + n_t - 1}"
        t_first = get_column_letter(2)
        t_last = get_column_letter(2 + n_t - 1)
        wrow = f"'Optimize Inputs'!${t_first}$5:${t_last}$5"
        murow = f"'Optimize Inputs'!${t_first}$6:${t_last}$6"
        # #region agent log
        _debug_log(
            "debug-pre-fix",
            "H4",
            "export_service.py:_add_portfolio_org_sheet",
            "portfolio_org_ranges_computed",
            {
                "nTickers": int(n_t),
                "nObs": int(n_obs),
                "weightsRow": wrow,
                "muRow": murow,
                "covTopLeft": cov_tl,
                "covBottomRight": cov_br,
                "orpCol": int(orp_col),
                "covRowStart": int(cov_r0),
            },
        )
        # #endregion

        ws.cell(
            row=1, column=1, value="Complete portfolio — inputs & CAL math"
        ).font = self.title_font
        ws.cell(row=2, column=1, value="Field").font = self.bold_font
        ws.cell(row=2, column=2, value="Modeled (formula)").font = self.bold_font
        ws.cell(row=2, column=4, value="Backend snapshot").font = self.bold_font

        r = 3
        r_rf = r
        ws.cell(row=r, column=1, value="Risk-free rate r_f (linked to Optimize Inputs)")
        ws.cell(row=r, column=2, value="='Optimize Inputs'!B2")
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=4, value=float(optimize.risk_free_rate)).number_format = "0.00%"
        r += 1
        r_a = r
        ws.cell(row=r, column=1, value="Risk aversion A (client, 1–10)")
        ws.cell(row=r, column=2, value=int(risk_profile.risk_aversion))
        r += 1
        r_tr = r
        ws.cell(row=r, column=1, value="Target annual return (optional)")
        tr_cell = ws.cell(row=r, column=2)
        if risk_profile.target_return is not None:
            tr_cell.value = float(risk_profile.target_return)
            tr_cell.number_format = "0.00%"
        r += 1
        r_al = r
        ws.cell(row=r, column=1, value="Allow leverage")
        ws.cell(row=r, column=2, value=bool(allow_leverage))
        r += 1
        ws.cell(row=r, column=1, value="allow_short (tangency vs QP ORP)")
        ws.cell(row=r, column=2, value=bool(allow_short))
        r += 1

        r_eorp = r
        ws.cell(row=r, column=1, value="E(r_ORP) = w'μ")
        ws.cell(row=r, column=2, value=f"={_excel_scalar(f'SUMPRODUCT({wrow},{murow})')}")
        ws.cell(row=r, column=2).number_format = "0.00%"
        ws.cell(row=r, column=4, value=float(optimize.orp.expected_return)).number_format = "0.00%"
        r += 1
        r_var = r
        ws.cell(row=r, column=1, value="Variance ORP = w'Σw (sample cov × k)")
        orp_col_letter = get_column_letter(orp_col)
        variance_inner = f"VAR.S('Optimize Inputs'!{orp_col_letter}{first_data_row}:{orp_col_letter}{last_data_row})"
        ws.cell(
            row=r,
            column=2,
            value=f"={_excel_scalar(variance_inner)}*'Optimize Inputs'!$B$1",
        )
        # #region agent log
        _debug_log(
            "debug-pre-fix",
            "H5",
            "export_service.py:_add_portfolio_org_sheet",
            "variance_formula_written",
            {
                "cell": f"B{r}",
                "formula": ws.cell(row=r, column=2).value,
                "orpColumn": orp_col_letter,
                "rowRange": f"{first_data_row}:{last_data_row}",
                "annualizationRef": "'Optimize Inputs'!$B$1",
            },
        )
        # #endregion
        ws.cell(row=r, column=2).number_format = "0.000000"
        ws.cell(row=r, column=4, value=float(optimize.orp.variance)).number_format = "0.000000"
        r += 1
        r_sig = r
        ws.cell(row=r, column=1, value="Std dev ORP")
        ws.cell(row=r, column=2, value=f"=SQRT(B{r_var})")
        ws.cell(row=r, column=2).number_format = "0.00%"
        self._percent_input_cell(ws, r, 4, float(optimize.orp.std_dev))
        r += 2

        r_rp = r
        ws.cell(row=r, column=1, value="Risk premium E(r_ORP) − r_f")
        c_rp = ws.cell(row=r, column=2, value=f"=B{r_eorp}-B{r_rf}")
        c_rp.number_format = "0.00%"
        r += 1
        r_yu = r
        ws.cell(row=r, column=1, value="y* from utility (risk premium / (A·Var))")
        c_yu = ws.cell(row=r, column=2)
        c_yu.value = f"=B{r_rp}/(B{r_a}*B{r_var})"
        c_yu.number_format = "0.00%"
        r += 1
        r_yfinal = r
        ws.cell(row=r, column=1, value="Final y* (target-return override + leverage rule)")
        yu = f"B{r_yu}"
        trr = f"B{r_tr}"
        rff = f"B{r_rf}"
        rpp = f"B{r_rp}"
        eor = f"B{r_eorp}"
        alw = f"B{r_al}"
        y_uncapped = (
            f"IF(AND(ISNUMBER({trr}),{trr}>{eor},{rpp}>0),"
            f"IF(({trr}-{rff})/{rpp}>{yu},({trr}-{rff})/{rpp},{yu}),{yu})"
        )
        yf = ws.cell(row=r, column=2)
        yf.value = f"=IF({alw},{y_uncapped},MIN({y_uncapped},1))"
        yf.number_format = "0.00%"
        r_yfinal_row = r_yfinal
        r += 1

        ws.cell(row=r, column=1, value="Weight in risk-free (1 − y*)")
        ws.cell(row=r, column=2, value=f"=1-B{r_yfinal_row}").number_format = "0.00%"
        r += 1
        ws.cell(row=r, column=1, value="E(r) complete portfolio")
        ws.cell(row=r, column=2, value=f"=B{r_rf}+B{r_yfinal_row}*B{r_rp}").number_format = "0.00%"
        r += 1
        ws.cell(row=r, column=1, value="σ complete portfolio")
        ws.cell(row=r, column=2, value=f"=B{r_yfinal_row}*B{r_sig}").number_format = "0.00%"
        r += 2

        ws.cell(row=r, column=1, value="ORP weights (Excel vs backend)").font = self.bold_font
        r += 1
        ws.cell(row=r, column=1, value="Ticker").font = self.bold_font
        ws.cell(row=r, column=2, value="w_ORP (Optimize Inputs)").font = self.bold_font
        ws.cell(row=r, column=3, value="w_complete").font = self.bold_font
        ws.cell(row=r, column=4, value="w_ORP backend").font = self.bold_font
        r += 1
        for i, stk in enumerate(optimize.stocks):
            t = stk.ticker
            oi_col = get_column_letter(2 + i)
            w_api = float(optimize.orp.weights.get(t, 0.0))
            ws.cell(row=r, column=1, value=t)
            ws.cell(row=r, column=2, value=f"='Optimize Inputs'!{oi_col}5")
            ws.cell(row=r, column=2).number_format = "0.00%"
            ws.cell(row=r, column=4, value=w_api).number_format = "0.00%"
            ws.cell(row=r, column=3, value=f"=B{r_yfinal_row}*B{r}").number_format = "0.00%"
            r += 1

        r += 1
        ws.cell(
            row=r, column=1, value="Correlation matrix (from aligned log returns)"
        ).font = self.bold_font
        r += 1
        ts = [s.ticker for s in optimize.stocks]
        for j, t in enumerate(ts, 2):
            ws.cell(row=r, column=j, value=t).font = self.bold_font
        r += 1
        for i, t_i in enumerate(ts):
            ws.cell(row=r, column=1, value=t_i).font = self.bold_font
            ci = get_column_letter(2 + i)
            for j, _t_j in enumerate(ts):
                cj = get_column_letter(2 + j)
                ws.cell(
                    row=r,
                    column=2 + j,
                    value=(
                        f"=CORREL('Optimize Inputs'!{ci}${first_data_row}:{ci}${last_data_row},"
                        f"'Optimize Inputs'!{cj}${first_data_row}:{cj}${last_data_row})"
                    ),
                ).number_format = "0.00"
            r += 1
